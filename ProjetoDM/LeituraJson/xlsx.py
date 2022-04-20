from openpyxl import Workbook, load_workbook

from openpyxl.utils import get_column_letter


class expXlsx:
    def export(id, nome, usuario, ip):
        #Abre o arquivo
        wb = load_workbook(R'./DB/patrimonio.xlsx')

        #Seleciona a planilha
        p1 = wb['Planilha1']
        p1 = wb.active

        #Estabelece e insere um cabeçalho
        header = ['ID', 'Nome', 'Usuário', 'Endereço']
        position = ['A', 'B', 'C', 'D']
        for i in range(len(header)):
            p1[f"{position[i]}1"] = f"{header[i]}"

        #Insere os dados na planilha
        data = [id, nome, usuario, ip]
        for i in range(len(data)):
            p1[f'{position[i]}{int(id)+1}'] = data[i]

        #Salva o arquivo
        try:
            wb.save(R'./DB/patrimonio.xlsx')
            wb.close()

        except:
            print('Não foi possível salvar o arquivo!')
